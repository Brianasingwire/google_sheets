'''Compare multiple Excel files'''

import openpyxl


def load_workbook(file_path):
    """Load an Excel workbook."""
    return openpyxl.load_workbook(file_path)


def compare_sheets(sheet1, sheet2):
    """Compare two sheets cell by cell."""

    max_row = max(sheet1.max_row, sheet2.max_row)
    max_column = max(sheet1.max_column, sheet2.max_column)

    differences = []

    for row in range(2, max_row + 1):
        for col in range(2, max_column + 1):
            cell1 = sheet1.cell(row=row, column=col)
            cell2 = sheet2.cell(row=row, column=col)

            value1 = cell1.value
            value2 = cell2.value

            if value1 != value2:
                differences.append({
                    'row': row,
                    'column': col,
                    'value1': value1,
                    'value2': value2
                })

    return differences


def save_differences(differences, output_file):
    """Save the differences to a new Excel sheet."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Differences"

    headers = ["Row", "Column", "Sheet1 Value", "Sheet2 Value"]
    sheet.append(headers)

    for difference in differences:
        row = [
            difference['row'],
            difference['column'],
            difference['value1'],
            difference['value2']
        ]
        sheet.append(row)

    wb.save(output_file)
    print(f"Differences saved to {output_file}")


def main():
    """Main function to compare two Excel sheets and save differences."""

    file1 = '../innovators.xlsx'
    file2 = '../python_data.xlsx'
    sheet_name1 = 'Data from csv'
    sheet_name2 = 'Python data'
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)

    sheet1 = wb1[sheet_name1]
    sheet2 = wb2[sheet_name2]

    differences = compare_sheets(sheet1, sheet2)

    output_file = 'differences.xlsx'
    save_differences(differences, output_file)


if __name__ == '__main__':
    main()
