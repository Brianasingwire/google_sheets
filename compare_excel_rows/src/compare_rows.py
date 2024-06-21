'''Comparing rows in an Excel file'''

import openpyxl


def load_workbook(filename):
    """Load the workbook and return the active worksheet."""

    workbook = openpyxl.load_workbook(filename)
    sheet_name = 'Data from csv'
    worksheet = workbook[sheet_name]
    return worksheet


def compare_rows(row1, row2):
    """Compare two rows and return True if they are identical, otherwise
    False."""
    return row1 == row2


def compare_all_rows(worksheet):
    """Compare all rows with each other and print the results."""

    rows = list(worksheet.iter_rows(values_only=True))
    for i, row1 in enumerate(rows, start=1):
        for j, row2 in enumerate(rows[i:], start=i+1):
            if compare_rows(row1, row2):
                print(f"Row {i} and Row {j} are identical")
            else:
                print(f"Row {i} and Row {j} are different")


def main():
    """_summary_
    """

    filename = '../innovators.xlsx'
    worksheet = load_workbook(filename)
    compare_all_rows(worksheet)


if __name__ == "__main__":
    main()
