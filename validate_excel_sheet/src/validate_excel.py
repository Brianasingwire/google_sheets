'''Validating an Excel sheet'''

from openpyxl import load_workbook


def check_excel_sheet(file_path):
    """_summary_

    Args:
        filepath (_type_): _description_
    """

    wb = load_workbook(file_path)
    ws = wb['Named Ranges']

    blank_rows = []
    invalid_rows = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        if all(cell is None for cell in row):
            blank_rows.append(row)

        elif any(cell is None for cell in row):
            invalid_rows.append(row)

    if blank_rows:
        print('Alert: Blank rows found in spreadsheet')
        for idx, row in enumerate(blank_rows, start=1):
            print(f'Blank row {idx}: {row}')
    else:
        print('No blank rows found in spreadsheet')

    if invalid_rows:
        print('Alert: Rows with invalid date found in spreadsheet.')
        for idx, row in enumerate(invalid_rows, start=1):
            print(f'Invalid data row {idx}: {row}')
    else:
        print('No invalid data found in spreadsheet')


if __name__ == '__main__':
    check_excel_sheet('../Excel 365  Advanced Example File.xlsx')
